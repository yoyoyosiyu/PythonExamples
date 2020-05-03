from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 42
ws.append([1,2,3])

import datetime

ws['A2']=datetime.datetime.now()

wb.save("sample.xlsx")
wb.close()

from openpyxl import load_workbook

wb = load_workbook(filename="sample.xlsx")
for sheet in wb:
    print(sheet)

sheet = wb['Sheet']
print(sheet['A1'].value)
